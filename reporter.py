"""
Generacion del informe Excel con formato profesional.

Escribe un libro de Excel con varias hojas (datos limpios, resumen,
agregaciones) aplicando estilos: cabeceras con color, anchos de columna
automaticos, formato de numero, filas con bandas, formato condicional y
un grafico de barras. Es la parte que convierte un dump de datos en algo
que el cliente percibe como "un informe".
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

# Paleta de la marca del informe. Cambia estos dos valores para
# adaptar el informe a los colores de un cliente concreto.
HEADER_FILL = "1F4E78"   # azul oscuro
BAND_FILL = "EAF1FB"     # azul muy claro para filas alternas

_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_FILL = PatternFill("solid", fgColor=HEADER_FILL)
_BAND_FILL = PatternFill("solid", fgColor=BAND_FILL)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_sheet(ws, df: pd.DataFrame, title: str | None = None) -> None:
    """Escribe un DataFrame en una hoja aplicando el estilo base."""
    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color=HEADER_FILL)
        start_row = 3

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.border = _BORDER
            is_header = r_idx == 0
            if is_header:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _CENTER
            else:
                # Bandas alternas para legibilidad.
                if r_idx % 2 == 0:
                    cell.fill = _BAND_FILL
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"

    _autosize_columns(ws, df, start_row)
    # Congela la fila de cabecera.
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _autosize_columns(ws, df: pd.DataFrame, start_row: int) -> None:
    """Ajusta el ancho de cada columna al contenido."""
    for i, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        data_len = df[col].apply(lambda v: len(str(v))).max() if len(df) else 0
        width = max(header_len, int(data_len or 0)) + 4
        ws.column_dimensions[get_column_letter(i)].width = min(width, 45)


def _add_bar_chart(ws, df: pd.DataFrame, start_row: int) -> None:
    """Anade un grafico de barras a una hoja de agregacion.

    Asume que la primera columna es la categoria y la segunda el valor.
    """
    if df.shape[1] < 2 or len(df) == 0:
        return

    data_start = start_row + 1  # fila siguiente a la cabecera
    data_end = data_start + len(df) - 1

    chart = BarChart()
    chart.type = "col"
    chart.title = df.columns[1]
    chart.height = 8
    chart.width = 16

    data = Reference(ws, min_col=2, min_row=start_row, max_row=data_end)
    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None

    anchor_col = get_column_letter(df.shape[1] + 2)
    ws.add_chart(chart, f"{anchor_col}{start_row}")


def build_report(
    output_path: str | Path,
    *,
    clean_data: pd.DataFrame,
    clean_report: pd.DataFrame,
    numeric_summary: pd.DataFrame,
    aggregation: pd.DataFrame | None = None,
    aggregation_title: str = "Agregacion",
) -> Path:
    """Construye y guarda el informe Excel completo.

    Crea las hojas: Resumen (limpieza + estadisticas), Datos limpios y,
    si se proporciona, una hoja de Agregacion con grafico.

    Returns:
        La ruta del archivo generado.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Hoja resumen: primero el log de limpieza, luego el resumen numerico.
        clean_report.to_excel(writer, sheet_name="Resumen", index=False, startrow=2)
        numeric_summary.to_excel(
            writer, sheet_name="Resumen", index=False,
            startrow=len(clean_report) + 6,
        )
        clean_data.to_excel(writer, sheet_name="Datos limpios", index=False)
        if aggregation is not None and not aggregation.empty:
            aggregation.to_excel(writer, sheet_name="Agregacion", index=False)

        wb = writer.book

        # Re-estilamos hoja a hoja sobre lo ya escrito.
        _restyle_summary(wb["Resumen"], clean_report, numeric_summary)
        _restyle_plain(wb["Datos limpios"], clean_data)
        if aggregation is not None and not aggregation.empty:
            agg_ws = wb["Agregacion"]
            _restyle_plain(agg_ws, aggregation, title=aggregation_title)
            _add_aggregation_extras(agg_ws, aggregation)

    return output_path


def _restyle_plain(ws, df: pd.DataFrame, title: str | None = None) -> None:
    """Limpia el contenido por defecto de pandas y reescribe con estilo."""
    ws.delete_rows(1, ws.max_row)
    _write_sheet(ws, df, title=title)


def _restyle_summary(ws, clean_report: pd.DataFrame, numeric_summary: pd.DataFrame) -> None:
    """Hoja resumen con dos bloques: limpieza y estadisticas."""
    ws.delete_rows(1, ws.max_row)
    ws.cell(row=1, column=1, value="INFORME DE DATOS").font = Font(
        bold=True, size=16, color=HEADER_FILL
    )

    # Bloque 1: que se ha limpiado.
    ws.cell(row=3, column=1, value="Proceso de limpieza").font = Font(bold=True, size=12)
    _write_block(ws, clean_report, start_row=4)

    # Bloque 2: estadisticas numericas.
    block2 = 4 + len(clean_report) + 3
    ws.cell(row=block2 - 1, column=1, value="Estadisticas numericas").font = Font(bold=True, size=12)
    _write_block(ws, numeric_summary, start_row=block2)
    _autosize_columns(ws, numeric_summary, block2)
    # La columna A debe acomodar las etiquetas largas del bloque de limpieza.
    label_width = max((len(str(v)) for v in clean_report.iloc[:, 0]), default=10) + 4
    current = ws.column_dimensions["A"].width or 0
    ws.column_dimensions["A"].width = max(current, label_width)


def _write_block(ws, df: pd.DataFrame, start_row: int) -> None:
    """Escribe un bloque de tabla con cabecera estilada en la fila indicada."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.border = _BORDER
            if r_idx == 0:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _CENTER
            elif isinstance(value, float):
                cell.number_format = "#,##0.00"


def _add_aggregation_extras(ws, df: pd.DataFrame) -> None:
    """Anade formato condicional (escala de color) y grafico a la agregacion."""
    # La hoja se reescribio con titulo => cabecera en la fila 3.
    header_row = 3
    data_start = header_row + 1
    data_end = data_start + len(df) - 1
    value_col = get_column_letter(2)

    ws.conditional_formatting.add(
        f"{value_col}{data_start}:{value_col}{data_end}",
        ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max", end_color="2E75B6",
        ),
    )
    _add_bar_chart(ws, df, header_row)
